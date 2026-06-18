"""vCenter (vSphere) read-only client — datastore 用量、snapshot 監控。

依賴：pip3 install pyvmomi
"""

import os
import ssl
import sys
import yaml
from datetime import datetime, timezone, timedelta
from pathlib import Path
from dotenv import load_dotenv

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))
load_dotenv(ROOT / ".env")
from tools.secrets import get as _secret

VC_HOST = os.getenv("VC_HOST", "10.11.1.15")
VC_USER = os.getenv("VC_USER", "infra.ro@uti.com")
VC_PASS = _secret("LDAP_SVC_PASS")   # infra.ro 與 infra_ldap 同密碼

_TZ8 = timezone(timedelta(hours=8))
_POWER = {"poweredOn": "On", "poweredOff": "Off", "suspended": "Suspended"}


def _connect():
    from pyVim.connect import SmartConnect
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    return SmartConnect(host=VC_HOST, user=VC_USER, pwd=VC_PASS, sslContext=ctx)


def _disconnect(si):
    from pyVim.connect import Disconnect
    Disconnect(si)


def _all_objects(si, vim_type):
    content = si.RetrieveContent()
    view = content.viewManager.CreateContainerView(content.rootFolder, [vim_type], True)
    try:
        return list(view.view)
    finally:
        view.Destroy()


# ── Datastore ──────────────────────────────────────────────

def get_datastores() -> list:
    """所有 datastore 用量，按使用率高→低。"""
    from pyVmomi import vim
    si = _connect()
    try:
        result = []
        for ds in _all_objects(si, vim.Datastore):
            s = ds.summary
            cap, free = s.capacity, s.freeSpace
            used = cap - free
            result.append({
                "name":        s.name,
                "type":        s.type,
                "accessible":  bool(s.accessible),
                "capacity_gb": round(cap / 1024**3, 1),
                "used_gb":     round(used / 1024**3, 1),
                "free_gb":     round(free / 1024**3, 1),
                "used_pct":    round(used / cap * 100, 1) if cap > 0 else 0,
            })
        return sorted(result, key=lambda x: -x["used_pct"])
    finally:
        _disconnect(si)


# ── Snapshot ───────────────────────────────────────────────

def _flatten_snapshots(tree, parent_path=""):
    out = []
    for s in tree:
        path = f"{parent_path}/{s.name}" if parent_path else s.name
        out.append({
            "name":        s.name,
            "path":        path,
            "id":          s.id,
            "created":     s.createTime,            # tz-aware datetime
            "description": s.description or "",
        })
        if s.childSnapshotList:
            out.extend(_flatten_snapshots(s.childSnapshotList, path))
    return out


def get_old_snapshots(days: int = 7) -> list:
    """
    回傳所有 snapshot 中，建立時間超過 N 天的。
    [{vm_name, vm_power, snap_name, snap_path, age_days, created (TZ+8), description}]
    """
    from pyVmomi import vim
    now = datetime.now(tz=timezone.utc)
    si = _connect()
    try:
        result = []
        for vm in _all_objects(si, vim.VirtualMachine):
            if not vm.snapshot:
                continue
            for snap in _flatten_snapshots(vm.snapshot.rootSnapshotList):
                age = (now - snap["created"]).days
                if age >= days:
                    result.append({
                        "vm_name":     vm.name,
                        "vm_power":    str(vm.runtime.powerState).replace("poweredO", ""),
                        "snap_name":   snap["name"],
                        "snap_path":   snap["path"],
                        "age_days":    age,
                        "created":     snap["created"].astimezone(_TZ8),
                        "description": snap["description"],
                    })
        return sorted(result, key=lambda x: -x["age_days"])
    finally:
        _disconnect(si)


# ── 報告 ──────────────────────────────────────────────────

def datastore_report(min_pct: float = 0) -> str:
    """Datastore 用量報告；min_pct 可過濾低使用率 datastore。"""
    ds_list = [d for d in get_datastores() if d["used_pct"] >= min_pct]
    lines = ["=" * 70, "  vCenter Datastore 用量", "=" * 70]
    for d in ds_list:
        icon = "🔴" if d["used_pct"] >= 85 else ("🟡" if d["used_pct"] >= 70 else "✅")
        lines.append(f"  {icon} {d['name'][:32]:32s} {d['type']:6s} "
                     f"{d['used_gb']:>8.0f} / {d['capacity_gb']:>8.0f} GB  ({d['used_pct']:>5.1f}%)")
    lines.append("=" * 70)
    return "\n".join(lines)


def snapshot_report(days: int = 7) -> str:
    """老舊 snapshot 報告。"""
    snaps = get_old_snapshots(days=days)
    lines = ["=" * 70, f"  vCenter 老舊 Snapshot（超過 {days} 天）", "=" * 70]
    if not snaps:
        lines.append(f"  ✅ 無超過 {days} 天的 snapshot")
    else:
        lines.append(f"  共 {len(snaps)} 筆")
        for s in snaps:
            icon = "🚨" if s["age_days"] >= 30 else ("🔴" if s["age_days"] >= 14 else "🟡")
            lines.append(f"  {icon} [{s['age_days']:3d}d] {s['vm_name'][:30]:30s} "
                         f"{s['snap_name'][:30]:30s} {s['created'].strftime('%Y-%m-%d %H:%M')}")
            if s["description"]:
                lines.append(f"        └─ {s['description'][:65]}")
    lines.append("=" * 70)
    return "\n".join(lines)


# ── VM 查詢 ─────────────────────────────────────────────────

def _vm_to_dict(vm) -> dict:
    """vim.VirtualMachine → 精簡 dict。"""
    ips = []
    if vm.guest and vm.guest.net:
        for nic in vm.guest.net:
            for ip in (nic.ipAddress or []):
                if ":" not in ip:   # 跳 IPv6
                    ips.append(ip)
    return {
        "name":       vm.name,
        "power":      _POWER.get(str(vm.runtime.powerState), str(vm.runtime.powerState)),
        "guest_os":   (vm.config.guestFullName if vm.config else ""),
        "ip":         ips[0] if ips else "",
        "ips":        ips,
        "host":       vm.runtime.host.name if vm.runtime.host else "",
        "datastores": [ds.name for ds in (vm.datastore or [])],
        "cpu":        vm.config.hardware.numCPU if vm.config else 0,
        "memory_mb":  vm.config.hardware.memoryMB if vm.config else 0,
        "tools":      str(vm.guest.toolsRunningStatus) if vm.guest else "",
    }


def get_all_vms() -> list:
    """所有 VM 簡明清單。"""
    from pyVmomi import vim
    si = _connect()
    try:
        return [_vm_to_dict(vm) for vm in _all_objects(si, vim.VirtualMachine)]
    finally:
        _disconnect(si)


def _find_vm(si, query: str):
    """name 完整 / 部分包含 / IP 命中。"""
    from pyVmomi import vim
    fallback = None
    for vm in _all_objects(si, vim.VirtualMachine):
        if vm.name == query:
            return vm
        if query in vm.name and fallback is None:
            fallback = vm
        if vm.guest and vm.guest.net:
            for nic in vm.guest.net:
                if query in (nic.ipAddress or []):
                    return vm
    return fallback


def get_vm(name_or_ip: str) -> dict:
    """單一 VM 詳查（disk / snapshot / boot / storage）。"""
    from pyVmomi import vim
    si = _connect()
    try:
        vm = _find_vm(si, name_or_ip)
        if not vm:
            raise ValueError(f"找不到 VM: {name_or_ip}")
        d = _vm_to_dict(vm)
        disks = []
        for dev in vm.config.hardware.device:
            if isinstance(dev, vim.vm.device.VirtualDisk):
                thin = getattr(dev.backing, "thinProvisioned", None)
                disks.append({
                    "label":   dev.deviceInfo.label,
                    "size_gb": round(dev.capacityInBytes / 1024**3, 1),
                    "mode":    "thin" if thin else ("thick" if thin is False else "?"),
                    "file":    dev.backing.fileName,
                })
        d["disks"] = disks
        d["snapshots"] = (_flatten_snapshots(vm.snapshot.rootSnapshotList)
                          if vm.snapshot else [])
        if vm.summary.storage:
            ss = vm.summary.storage
            d["storage"] = {
                "committed_gb":   round(ss.committed   / 1024**3, 1),
                "uncommitted_gb": round(ss.uncommitted / 1024**3, 1),
                "unshared_gb":    round(ss.unshared    / 1024**3, 1),
            }
        if vm.runtime.bootTime:
            d["boot_time"]   = vm.runtime.bootTime.astimezone(_TZ8)
            d["uptime_days"] = (datetime.now(tz=timezone.utc) - vm.runtime.bootTime).days
        return d
    finally:
        _disconnect(si)


def vm_report(name_or_ip: str) -> str:
    d = get_vm(name_or_ip)
    lines = ["=" * 70, f"  VM: {d['name']}", "=" * 70]
    lines.append(f"  Power     : {d['power']}")
    lines.append(f"  Guest OS  : {d['guest_os']}")
    lines.append(f"  IP        : {', '.join(d['ips']) if d['ips'] else '(none)'}")
    lines.append(f"  Host      : {d['host']}")
    lines.append(f"  Datastore : {', '.join(d['datastores'])}")
    lines.append(f"  CPU / RAM : {d['cpu']} vCPU / {d['memory_mb']/1024:.0f} GB")
    lines.append(f"  Tools     : {d['tools']}")
    if "boot_time" in d:
        lines.append(f"  Boot Time : {d['boot_time'].strftime('%Y-%m-%d %H:%M')} ({d['uptime_days']} 天)")
    if "storage" in d:
        s = d["storage"]
        lines.append(f"  Storage   : committed {s['committed_gb']:.1f} GB  unshared {s['unshared_gb']:.1f} GB  uncommitted {s['uncommitted_gb']:.1f} GB")
    lines.append("\n  Disks:")
    for disk in d["disks"]:
        lines.append(f"    {disk['label']}: {disk['size_gb']:.0f} GB ({disk['mode']})  {disk['file']}")
    if d["snapshots"]:
        lines.append(f"\n  Snapshots ({len(d['snapshots'])}):")
        for s in d["snapshots"]:
            t = s["created"].astimezone(_TZ8).strftime("%Y-%m-%d %H:%M")
            lines.append(f"    - {s['name']}  created={t}")
    else:
        lines.append("\n  Snapshots : (none)")
    lines.append("=" * 70)
    return "\n".join(lines)


# ── 對照盤點 ────────────────────────────────────────────────

def audit_coverage() -> dict:
    """vCenter VM ↔ inventory.yaml ↔ Zabbix ↔ PPDM 三方對照。"""
    vc_vms = get_all_vms()

    inv_ips = {}
    try:
        with open(ROOT / "inventory.yaml") as f:
            inv = yaml.safe_load(f)
        for label, n in (inv.get("nodes") or {}).items():
            if n and n.get("ip"):
                inv_ips[n["ip"]] = label
    except Exception:
        pass

    zb_names, zb_ips, zb_err = set(), set(), None
    try:
        from tools import zabbix_runner as zb
        t = zb.login()
        try:
            zhosts = zb._call("host.get", {
                "output": ["name", "host"],
                "selectInterfaces": ["ip"],
                "monitored_hosts": True,
            }, t)
        finally:
            zb.logout(t)
        for h in zhosts:
            zb_names.add(h.get("name", ""))
            zb_names.add(h.get("host", ""))
            for i in h.get("interfaces", []):
                if i.get("ip"):
                    zb_ips.add(i["ip"])
    except Exception as e:
        zb_err = str(e)

    ppdm_protected, ppdm_err = set(), None
    try:
        from tools import ppdm_runner as ppdm
        tk = ppdm.login()
        try:
            assets = ppdm.get_assets(tk, asset_type="VMWARE_VIRTUAL_MACHINE")
        finally:
            ppdm.logout(tk)
        for a in assets:
            if a.get("protectionStatus") != "UNPROTECTED":
                ppdm_protected.add(a.get("name", ""))
    except Exception as e:
        ppdm_err = str(e)

    rows = []
    for vm in vc_vms:
        in_inv  = bool(vm["ip"]) and vm["ip"] in inv_ips
        in_zb   = (vm["name"] in zb_names) or (bool(vm["ip"]) and vm["ip"] in zb_ips)
        in_ppdm = vm["name"] in ppdm_protected
        rows.append({**vm, "in_inventory": in_inv, "in_zabbix": in_zb, "in_ppdm": in_ppdm})

    on = [r for r in rows if r["power"] == "On"]
    return {
        "vms":              rows,
        "total":            len(rows),
        "powered_on":       on,
        "not_in_inventory": [r for r in rows if not r["in_inventory"]],
        "not_in_zabbix":    [r for r in on   if not r["in_zabbix"]],
        "not_in_ppdm":      [r for r in on   if not r["in_ppdm"]],
        "errors":           {k: v for k, v in {"zabbix": zb_err, "ppdm": ppdm_err}.items() if v},
    }


def audit_report() -> str:
    a = audit_coverage()
    lines = ["=" * 70, "  vCenter VM ↔ inventory / Zabbix / PPDM 對照盤點", "=" * 70]
    lines.append(f"\n  vCenter VM 總數: {a['total']}（poweredOn: {len(a['powered_on'])}）")
    if a["errors"]:
        lines.append("  ⚠️  資料源失敗: " + ", ".join(f"{k}={v[:60]}" for k, v in a["errors"].items()))

    if a["not_in_inventory"]:
        lines.append(f"\n【不在 inventory.yaml】{len(a['not_in_inventory'])} 台")
        for r in a["not_in_inventory"]:
            lines.append(f"  ⬜ {r['name'][:38]:38s} {r['ip'][:15]:15s} {r['power']:3s}")

    if a["not_in_zabbix"]:
        lines.append(f"\n【未被 Zabbix 監控（poweredOn）】{len(a['not_in_zabbix'])} 台")
        for r in a["not_in_zabbix"]:
            lines.append(f"  🔴 {r['name'][:38]:38s} {r['ip'][:15]:15s}")

    if a["not_in_ppdm"]:
        lines.append(f"\n【未被 PPDM 保護（poweredOn）】{len(a['not_in_ppdm'])} 台")
        for r in a["not_in_ppdm"]:
            lines.append(f"  🔴 {r['name'][:38]:38s} {r['ip'][:15]:15s}")

    lines.append("\n" + "=" * 70)
    return "\n".join(lines)


# ── ESXi Host 健康 ──────────────────────────────────────────

def get_hosts() -> list:
    """所有 ESXi host 容量、用量、over-commit。"""
    from pyVmomi import vim
    si = _connect()
    try:
        result = []
        for h in _all_objects(si, vim.HostSystem):
            hw = h.summary.hardware
            qs = h.summary.quickStats
            cpu_total_mhz = hw.cpuMhz * hw.numCpuCores
            cpu_used_mhz  = qs.overallCpuUsage or 0
            mem_total_mb  = hw.memorySize / 1024**2
            mem_used_mb   = qs.overallMemoryUsage or 0

            alloc_vcpu, alloc_mem = 0, 0
            for vm in h.vm:
                if vm.config:
                    alloc_vcpu += vm.config.hardware.numCPU
                    alloc_mem  += vm.config.hardware.memoryMB

            result.append({
                "name":            h.name,
                "version":         h.summary.config.product.fullName,
                "connection":      str(h.runtime.connectionState),
                "in_maint":        bool(h.runtime.inMaintenanceMode),
                "cpu_cores":       hw.numCpuCores,
                "cpu_threads":     hw.numCpuThreads,
                "cpu_used_pct":    round(cpu_used_mhz / cpu_total_mhz * 100, 1) if cpu_total_mhz else 0,
                "mem_total_gb":    round(mem_total_mb / 1024, 1),
                "mem_used_gb":     round(mem_used_mb / 1024, 1),
                "mem_used_pct":    round(mem_used_mb / mem_total_mb * 100, 1) if mem_total_mb else 0,
                "vm_count":        len(h.vm),
                "alloc_vcpu":      alloc_vcpu,
                "alloc_mem_gb":    round(alloc_mem / 1024, 1),
                "vcpu_overcommit": round(alloc_vcpu / hw.numCpuCores, 2) if hw.numCpuCores else 0,
                "mem_overcommit":  round(alloc_mem / mem_total_mb, 2) if mem_total_mb else 0,
            })
        return sorted(result, key=lambda x: x["name"])
    finally:
        _disconnect(si)


def host_report() -> str:
    hosts = get_hosts()
    lines = ["=" * 92, "  ESXi Host 健康", "=" * 92]
    lines.append(f"  {'狀態':4s}{'Host':27s} {'CPU%':>5s} {'MEM%':>5s} {'VMs':>4s}  "
                 f"{'vCPU(alloc/core, ovc)':>23s}  {'MEM(alloc/total GB, ovc)':>26s}")
    for h in hosts:
        flag = "🔧" if h["in_maint"] else ("✅" if h["connection"] == "connected" else "❌")
        lines.append(f"  {flag} {h['name'][:27]:27s} {h['cpu_used_pct']:>5.1f} {h['mem_used_pct']:>5.1f} "
                     f"{h['vm_count']:>4d}  {h['alloc_vcpu']:>4d}/{h['cpu_cores']:<3d} ({h['vcpu_overcommit']:.1f}x)"
                     f"      {h['alloc_mem_gb']:>6.0f}/{h['mem_total_gb']:.0f} ({h['mem_overcommit']:.2f}x)")
    lines.append("=" * 92)
    return "\n".join(lines)


# ── CPU 時間序列 ────────────────────────────────────────────

def get_cpu_timeseries(targets: list, start: "datetime", interval_id: int = 7200) -> dict:
    """
    查多個 entity（ESXi host 或 VM）的 CPU 時間序列。
    targets: [{"label": str, "ip": str, "type": "host"|"vm"}]
    回傳：{label: [(datetime, pct), ...]}
    """
    from pyVmomi import vim
    now_utc = datetime.now(tz=timezone.utc)
    si = _connect()
    try:
        pm = si.content.perfManager
        cpu_counter_id = None
        for c in pm.perfCounter:
            if c.groupInfo.key == "cpu" and c.nameInfo.key == "usage" and str(c.rollupType) == "average":
                cpu_counter_id = c.key
                break

        metric = vim.PerformanceManager.MetricId(counterId=cpu_counter_id, instance="")

        entity_map = {}
        for t in targets:
            if t["type"] == "host":
                for h in _all_objects(si, vim.HostSystem):
                    if h.name == t["ip"]:
                        entity_map[t["label"]] = h
                        break
            else:
                for vm in _all_objects(si, vim.VirtualMachine):
                    if vm.guest and vm.guest.net:
                        for nic in vm.guest.net:
                            if t["ip"] in (nic.ipAddress or []):
                                entity_map[t["label"]] = vm
                                break

        queries = [
            vim.PerformanceManager.QuerySpec(
                entity=entity_map[label],
                metricId=[metric],
                intervalId=interval_id,
                startTime=start.astimezone(timezone.utc),
                endTime=now_utc,
            )
            for label in entity_map
        ]
        results = pm.QueryPerf(querySpec=queries)

        out = {}
        labels = list(entity_map.keys())
        for i, r in enumerate(results):
            if not r.value: continue
            label = labels[i]
            out[label] = [
                (ts.timestamp.astimezone(_TZ8), v / 100)
                for ts, v in zip(r.sampleInfo, r.value[0].value)
            ]
        return out
    finally:
        _disconnect(si)


def cpu_timeseries_report(targets: list, start: "datetime", bucket_hours: int = 2) -> str:
    """
    用 5min 樣本抓資料（避免 2h rollup 延遲），然後以 bucket_hours 為單位聚合顯示。
    """
    data = get_cpu_timeseries(targets, start, interval_id=300)
    if not data:
        return "無資料"

    # bucket: 把每個 timestamp 對齊到 bucket_hours 的邊界（取 avg）
    from collections import defaultdict
    import math

    def _bucket_key(dt: "datetime") -> "datetime":
        h = (dt.hour // bucket_hours) * bucket_hours
        return dt.replace(hour=h, minute=0, second=0, microsecond=0)

    bucketed = {}
    for lb, rows in data.items():
        buckets: dict = defaultdict(list)
        for t, v in rows:
            buckets[_bucket_key(t)].append(v)
        bucketed[lb] = {k: sum(vs) / len(vs) for k, vs in buckets.items()}

    labels = list(data.keys())
    all_times = sorted({t for rows in bucketed.values() for t in rows})

    header = f"  {'時間':16s}" + "".join(f" {lb[:12]:>12s}" for lb in labels)
    sep    = "  " + "─" * (16 + 13 * len(labels))
    lines  = ["=" * (18 + 13 * len(labels)),
              f"  CPU Loading 時間序列（每 {bucket_hours}h 平均）",
              "=" * (18 + 13 * len(labels)), header, sep]

    for t in all_times:
        row = f"  {t.strftime('%m-%d %H:%M'):16s}"
        for lb in labels:
            v = bucketed[lb].get(t)
            if v is None:
                row += f" {'─':>12s}"
            else:
                flag = "🔴" if v >= 80 else ("🟡" if v >= 50 else "  ")
                row += f" {flag}{v:>8.1f}%"
        lines.append(row)

    lines.append(sep)
    for lb in labels:
        vals = list(bucketed[lb].values())
        if vals:
            lines.append(f"  {lb[:16]:16s} avg={sum(vals)/len(vals):.1f}%  peak={max(vals):.1f}%")
    lines.append("=" * (18 + 13 * len(labels)))
    return "\n".join(lines)


def get_host_cpu_timeseries(host_ip: str, start: "datetime") -> dict:
    """
    查指定 ESXi host 本身 + 其上所有 VM 的 CPU 5min 時間序列。
    回傳：{label: [(datetime, pct), ...]}
    label 格式：host 用 host_ip，VM 用 vm.name
    """
    from pyVmomi import vim
    now_utc = datetime.now(tz=timezone.utc)
    si = _connect()
    try:
        pm = si.content.perfManager
        cpu_counter_id = None
        for c in pm.perfCounter:
            if c.groupInfo.key == "cpu" and c.nameInfo.key == "usage" and str(c.rollupType) == "average":
                cpu_counter_id = c.key
                break

        metric = vim.PerformanceManager.MetricId(counterId=cpu_counter_id, instance="")

        target_host = None
        for h in _all_objects(si, vim.HostSystem):
            if h.name == host_ip:
                target_host = h
                break
        if not target_host:
            raise ValueError(f"找不到 ESXi host: {host_ip}")

        entities = [(host_ip, target_host)] + [
            (vm.name, vm) for vm in target_host.vm if vm.config
        ]

        queries = [
            vim.PerformanceManager.QuerySpec(
                entity=obj,
                metricId=[metric],
                intervalId=300,
                startTime=start.astimezone(timezone.utc),
                endTime=now_utc,
            )
            for _, obj in entities
        ]
        results = pm.QueryPerf(querySpec=queries)

        out = {}
        for i, r in enumerate(results):
            if not r.value:
                continue
            label = entities[i][0]
            out[label] = [
                (ts.timestamp.astimezone(_TZ8), v / 100)
                for ts, v in zip(r.sampleInfo, r.value[0].value)
            ]
        return out
    finally:
        _disconnect(si)


def export_cpu_excel(host_ip: str, start: "datetime", bucket_hours: int = 1, out_path: str = None) -> str:
    """
    查 ESXi host + 所有 VM 的 CPU 時間序列，匯出為 Excel。
    回傳儲存路徑。
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from collections import defaultdict

    data = get_host_cpu_timeseries(host_ip, start)
    if not data:
        raise RuntimeError("無資料")

    def _bucket_key(dt):
        h = (dt.hour // bucket_hours) * bucket_hours
        return dt.replace(hour=h, minute=0, second=0, microsecond=0)

    bucketed = {}
    for lb, rows in data.items():
        buckets = defaultdict(list)
        for t, v in rows:
            buckets[_bucket_key(t)].append(v)
        bucketed[lb] = {k: sum(vs) / len(vs) for k, vs in buckets.items()}

    all_times = sorted({t for rows in bucketed.values() for t in rows})
    # host first, then VMs sorted by name
    labels = [host_ip] + sorted(k for k in bucketed if k != host_ip)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CPU Loading"

    fill_red    = PatternFill("solid", fgColor="FF4444")
    fill_yellow = PatternFill("solid", fgColor="FFCC00")
    fill_header = PatternFill("solid", fgColor="2E75B6")
    font_white  = Font(color="FFFFFF", bold=True)
    font_bold   = Font(bold=True)
    thin        = Side(style="thin", color="CCCCCC")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    center      = Alignment(horizontal="center")

    # header row
    ws.cell(1, 1, "時間").fill = fill_header
    ws.cell(1, 1).font = font_white
    ws.cell(1, 1).alignment = center
    ws.column_dimensions["A"].width = 18
    for col, lb in enumerate(labels, 2):
        c = ws.cell(1, col, lb)
        c.fill = fill_header
        c.font = font_white
        c.alignment = center
        col_letter = openpyxl.utils.get_column_letter(col)
        ws.column_dimensions[col_letter].width = max(18, len(lb) + 2)

    # data rows
    for row, t in enumerate(all_times, 2):
        ws.cell(row, 1, t.strftime("%m-%d %H:%M")).alignment = center
        ws.cell(row, 1).border = border
        for col, lb in enumerate(labels, 2):
            v = bucketed[lb].get(t)
            c = ws.cell(row, col)
            c.border = border
            c.alignment = center
            if v is None:
                c.value = "─"
            else:
                c.value = round(v, 1)
                c.number_format = "0.0"
                if v >= 80:
                    c.fill = fill_red
                    c.font = Font(bold=True)
                elif v >= 50:
                    c.fill = fill_yellow

    ws.freeze_panes = "B2"

    if out_path is None:
        from datetime import datetime as _dt
        ts = _dt.now().strftime("%Y%m%d_%H%M")
        out_path = f"/Users/hank_lin/Documents/cpu_loading_{host_ip.replace('.','_')}_{ts}.xlsx"

    wb.save(out_path)
    return out_path


# ── VM CPU 歷史 ─────────────────────────────────────────────

def get_cpu_usage(host_ip: str, days: int = 30) -> list:
    """
    指定 ESXi host 上所有 VM 過去 N 天的平均 CPU%。
    intervalId 自動選擇：days<=1 → 300s；days<=32 → 7200s；else → 86400s。
    回傳：[{vm_name, vcpu, avg_pct, max_pct, samples}]，按 avg_pct 高→低。
    """
    from pyVmomi import vim
    now = datetime.now(tz=timezone.utc)

    if days <= 1:
        interval_id = 300
    elif days <= 32:
        interval_id = 7200
    else:
        interval_id = 86400

    si = _connect()
    try:
        pm = si.content.perfManager

        cpu_counter_id = None
        for c in pm.perfCounter:
            if (c.groupInfo.key == "cpu" and c.nameInfo.key == "usage"
                    and str(c.rollupType) == "average"):
                cpu_counter_id = c.key
                break
        if cpu_counter_id is None:
            raise RuntimeError("找不到 cpu.usage.average 計數器")

        # 找目標 host
        target_host = None
        for h in _all_objects(si, vim.HostSystem):
            if h.name == host_ip:
                target_host = h
                break
        if not target_host:
            raise ValueError(f"找不到 ESXi host: {host_ip}")

        vms = [vm for vm in target_host.vm if vm.config]
        if not vms:
            return []

        metric = vim.PerformanceManager.MetricId(counterId=cpu_counter_id, instance="")
        queries = [
            vim.PerformanceManager.QuerySpec(
                entity=vm,
                metricId=[metric],
                intervalId=interval_id,
                startTime=now - timedelta(days=days),
                endTime=now,
            )
            for vm in vms
        ]
        results = pm.QueryPerf(querySpec=queries)

        vm_map = {vm._moId: vm for vm in vms}
        result = []
        for r in results:
            vm = vm_map.get(r.entity._moId)
            if not vm:
                continue
            vals = [v for series in r.value for v in series.value if v >= 0]
            if not vals:
                avg, peak = 0.0, 0.0
            else:
                avg  = round(sum(vals) / len(vals) / 100, 2)
                peak = round(max(vals) / 100, 2)
            result.append({
                "vm_name": vm.name,
                "vcpu":    vm.config.hardware.numCPU,
                "avg_pct": avg,
                "max_pct": peak,
                "samples": len(vals),
            })

        return sorted(result, key=lambda x: -x["avg_pct"])
    finally:
        _disconnect(si)


def cpu_report(host_ip: str, days: int = 30) -> str:
    """VM CPU 平均使用率報告。"""
    rows = get_cpu_usage(host_ip=host_ip, days=days)
    lines = ["=" * 75,
             f"  ESXi {host_ip} — VM CPU 平均 Loading（過去 {days} 天）",
             "=" * 75,
             f"  {'VM Name':38s} {'vCPU':>4s} {'Avg%':>6s} {'Max%':>6s} {'樣本數':>6s}"]
    for r in rows:
        flag = "🔴" if r["avg_pct"] >= 80 else ("🟡" if r["avg_pct"] >= 50 else "  ")
        lines.append(f"  {flag} {r['vm_name'][:36]:36s} {r['vcpu']:>4d} "
                     f"{r['avg_pct']:>6.2f} {r['max_pct']:>6.2f} {r['samples']:>6d}")
    lines.append("=" * 75)
    return "\n".join(lines)


# ── 觸發中 Alarm ────────────────────────────────────────────

def get_active_alarms() -> list:
    """所有觸發中（紅 / 黃）的 alarm。"""
    from pyVmomi import vim
    si = _connect()
    try:
        result = []
        for cls in [vim.ClusterComputeResource, vim.HostSystem,
                    vim.VirtualMachine, vim.Datastore]:
            for ent in _all_objects(si, cls):
                for ta in (ent.triggeredAlarmState or []):
                    if ta.overallStatus not in ("red", "yellow"):
                        continue
                    try:
                        alarm_name = ta.alarm.info.name
                    except Exception:
                        alarm_name = "?"
                    result.append({
                        "entity":       ent.name,
                        "entity_type":  cls.__name__.replace("ComputeResource", ""),
                        "alarm":        alarm_name,
                        "severity":     ta.overallStatus,
                        "time":         ta.time.astimezone(_TZ8) if ta.time else None,
                        "acknowledged": bool(ta.acknowledged),
                    })
        return sorted(result, key=lambda x: (x["severity"] != "red",
                                              x["time"] or datetime.min.replace(tzinfo=_TZ8)))
    finally:
        _disconnect(si)


def alarm_report() -> str:
    alarms = get_active_alarms()
    lines = ["=" * 80, "  vCenter 觸發中 Alarm", "=" * 80]
    if not alarms:
        lines.append("  ✅ 無觸發中的 alarm")
    else:
        lines.append(f"  共 {len(alarms)} 筆")
        for a in alarms:
            icon = "🔴" if a["severity"] == "red" else "🟡"
            ack  = " (ack)" if a["acknowledged"] else ""
            t    = a["time"].strftime("%Y-%m-%d %H:%M") if a["time"] else "?"
            lines.append(f"  {icon} [{t}] {a['entity_type']:8s} {a['entity'][:28]:28s} {a['alarm'][:35]}{ack}")
    lines.append("=" * 80)
    return "\n".join(lines)


# ── 24h Events 摘要 ─────────────────────────────────────────

_KEY_EVENT_TYPES = [
    "VmMigratedEvent", "VmPoweredOnEvent", "VmPoweredOffEvent",
    "VmReconfiguredEvent", "VmRemovedEvent", "VmCreatedEvent",
    "VmGuestRebootEvent", "VmRebootGuestEvent",
    "TaskEvent", "AlarmStatusChangedEvent",
    "HostConnectionLostEvent", "HostConnectedEvent",
]


def get_recent_events(hours: int = 24, types: list = None) -> list:
    """過去 N 小時的關鍵事件。"""
    from pyVmomi import vim
    si = _connect()
    try:
        content = si.RetrieveContent()
        spec = vim.event.EventFilterSpec()
        spec.eventTypeId = types or _KEY_EVENT_TYPES
        tf = vim.event.EventFilterSpec.ByTime()
        tf.beginTime = datetime.now(tz=timezone.utc) - timedelta(hours=hours)
        spec.time = tf
        events = content.eventManager.QueryEvents(spec)
        result = []
        for e in events:
            result.append({
                "type":    type(e).__name__.replace("vim.event.", ""),
                "time":    e.createdTime.astimezone(_TZ8) if e.createdTime else None,
                "user":    e.userName or "",
                "vm":      e.vm.name if getattr(e, "vm", None) and e.vm else "",
                "host":    e.host.name if getattr(e, "host", None) and e.host else "",
                "message": (e.fullFormattedMessage or "")[:200],
            })
        return result
    finally:
        _disconnect(si)


def events_report(hours: int = 24) -> str:
    events = get_recent_events(hours=hours)
    lines = ["=" * 80, f"  vCenter Events 摘要（過去 {hours} 小時）", "=" * 80]

    by_type = {}
    for e in events:
        by_type[e["type"]] = by_type.get(e["type"], 0) + 1
    if by_type:
        lines.append(f"\n【統計】共 {len(events)} 筆")
        for t, c in sorted(by_type.items(), key=lambda x: -x[1]):
            lines.append(f"  {c:>5d}  {t}")

    KEY = {"VmMigratedEvent", "VmPoweredOffEvent", "HostConnectionLostEvent",
           "AlarmStatusChangedEvent", "VmRemovedEvent"}
    important = [e for e in events if e["type"] in KEY]
    if important:
        lines.append(f"\n【重要事件】{len(important)} 筆（顯示前 50）")
        for e in important[:50]:
            t   = e["time"].strftime("%m-%d %H:%M") if e["time"] else "?"
            who = f" by {e['user']}" if e["user"] else ""
            lines.append(f"  [{t}] {e['type']:25s} {e['vm'][:22]:22s} {e['message'][:55]}{who}")
    else:
        lines.append(f"\n  ✅ 無 vMotion / power off / 主機斷線等重要事件")

    lines.append("\n" + "=" * 80)
    return "\n".join(lines)


# ── Datastore 詳查 ──────────────────────────────────────────

def get_datastore_vms(name: str) -> dict:
    """指定 datastore 上的 VM 清單與用量。"""
    from pyVmomi import vim
    si = _connect()
    try:
        target = None
        for ds in _all_objects(si, vim.Datastore):
            if ds.name == name:
                target = ds; break
        if not target:
            raise ValueError(f"找不到 datastore: {name}")

        s = target.summary
        info = {
            "name":        s.name,
            "type":        s.type,
            "capacity_gb": round(s.capacity / 1024**3, 1),
            "used_gb":     round((s.capacity - s.freeSpace) / 1024**3, 1),
            "free_gb":     round(s.freeSpace / 1024**3, 1),
            "used_pct":    round((s.capacity - s.freeSpace) / s.capacity * 100, 1) if s.capacity > 0 else 0,
            "vms":         [],
        }
        for vm in target.vm:
            # 取「該 VM 在目標 datastore 上」的占用，而非全 VM 總和
            pdu = None
            if vm.storage and vm.storage.perDatastoreUsage:
                for u in vm.storage.perDatastoreUsage:
                    if u.datastore == target:
                        pdu = u; break
            info["vms"].append({
                "name":           vm.name,
                "power":          _POWER.get(str(vm.runtime.powerState), str(vm.runtime.powerState)),
                "committed_gb":   round(pdu.committed   / 1024**3, 1) if pdu else 0,
                "uncommitted_gb": round(pdu.uncommitted / 1024**3, 1) if pdu else 0,
                "unshared_gb":    round(pdu.unshared    / 1024**3, 1) if pdu else 0,
            })
        info["vms"].sort(key=lambda x: -x["committed_gb"])
        return info
    finally:
        _disconnect(si)


def datastore_detail_report(name: str, top: int = 30) -> str:
    info = get_datastore_vms(name)
    lines = ["=" * 88, f"  Datastore 詳查: {info['name']} ({info['type']})", "=" * 88]
    lines.append(f"  容量: {info['used_gb']:.0f} / {info['capacity_gb']:.0f} GB ({info['used_pct']:.1f}%)   剩餘 {info['free_gb']:.0f} GB")
    lines.append(f"\n【VM 清單，按 committed 大→小，前 {top} 筆】共 {len(info['vms'])} 台")
    for v in info["vms"][:top]:
        lines.append(f"  {v['power']:3s} {v['name'][:38]:38s} "
                     f"committed={v['committed_gb']:>7.1f} GB  "
                     f"unshared={v['unshared_gb']:>7.1f} GB  "
                     f"uncommitted={v['uncommitted_gb']:>6.1f} GB")
    lines.append("=" * 88)
    return "\n".join(lines)


# ── VMFS UUID 查詢 ──────────────────────────────────────────

def get_vmfs_uuids() -> list:
    """所有已掛載 VMFS datastore 的 UUID。"""
    from pyVmomi import vim
    si = _connect()
    try:
        result = []
        for ds in _all_objects(si, vim.Datastore):
            info = ds.info
            if not (hasattr(info, "vmfs") and info.vmfs):
                continue
            hosts = [m.key.name for m in (ds.host or [])]
            result.append({
                "name":  ds.summary.name,
                "uuid":  info.vmfs.uuid,
                "hosts": hosts,
            })
        return sorted(result, key=lambda x: x["name"])
    finally:
        _disconnect(si)


def vmfs_uuid_report() -> str:
    rows = get_vmfs_uuids()
    lines = ["=" * 90, "  VMFS UUID 清單（已掛載）", "=" * 90,
             f"  {'Datastore':40s} {'VMFS UUID':36s} 掛載主機"]
    for r in rows:
        hosts = ", ".join(r["hosts"])
        lines.append(f"  {r['name'][:40]:40s} {r['uuid']:36s} {hosts}")
    lines.append("=" * 90)
    return "\n".join(lines)


# ── CLI ──────────────────────────────────────────────────

_USAGE = """Usage: vsphere_runner.py <subcommand> [args]
  ds [name]               Datastore 用量；給 name 列上面 VM
  snap [days=7]           超過 N 天的 snapshot
  vm <name|ip>            VM 詳查
  audit                   vCenter ↔ inventory ↔ Zabbix ↔ PPDM 對照盤點
  hosts                   ESXi host 健康
  alarms                  觸發中 alarm
  events [hours=24]       過去 N 小時關鍵事件
  cpu <host_ip> [days=30] 指定 ESXi 上所有 VM CPU 平均 loading
  vmfs-uuid               所有已掛載 VMFS datastore 的 UUID
  all [days=7]            datastore + snapshot
"""

if __name__ == "__main__":
    cmd = sys.argv[1] if len(sys.argv) > 1 else "all"

    if cmd in ("ds", "datastore", "datastores"):
        if len(sys.argv) > 2:
            print(datastore_detail_report(sys.argv[2]))
        else:
            print(datastore_report())
    elif cmd in ("snap", "snapshot", "snapshots"):
        days = int(sys.argv[2]) if len(sys.argv) > 2 else 7
        print(snapshot_report(days=days))
    elif cmd in ("vm", "v"):
        if len(sys.argv) < 3:
            print(_USAGE); sys.exit(1)
        print(vm_report(sys.argv[2]))
    elif cmd in ("audit", "coverage"):
        print(audit_report())
    elif cmd in ("hosts", "host", "esxi"):
        print(host_report())
    elif cmd in ("alarms", "alarm"):
        print(alarm_report())
    elif cmd in ("events", "event"):
        hours = int(sys.argv[2]) if len(sys.argv) > 2 else 24
        print(events_report(hours=hours))
    elif cmd in ("cpu",):
        if len(sys.argv) < 3:
            print(_USAGE); sys.exit(1)
        host_ip = sys.argv[2]
        days = int(sys.argv[3]) if len(sys.argv) > 3 else 30
        print(cpu_report(host_ip=host_ip, days=days))
    elif cmd in ("vmfs-uuid", "vmfs_uuid", "uuid"):
        print(vmfs_uuid_report())
    elif cmd == "all":
        days = int(sys.argv[2]) if len(sys.argv) > 2 else 7
        print(datastore_report()); print()
        print(snapshot_report(days=days))
    else:
        print(_USAGE); sys.exit(1)
